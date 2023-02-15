KMD_WIDTH=0.0095
MZ_WIDTH=0.1
PRECISION = 2
PRECISION_APPEAR_IN_ALL_SAMPLES = 3

## for every excel file
DIR_SAMPLE                  = "0.sample"
DIR_BLANK                   = "1.blank"
DIR_FILTERED_SAMPLE         = "2.sample filtered with blank"
DIR_FILTRERD_INTENSITY_LIST = "3.sample filtered with intensitys"
DIR_RESULT                  = "4.result"
INTENSITY_LIST = [5000,10000,20000,50000,100000,200000,500000]

import time,os,sys,re,shutil,openpyxl
from operator import itemgetter
from pymongo import MongoClient
from bson.objectid import ObjectId
import pyzipper
HEAD_TITLE = ("m/z", "intensity", "kmd", "possible PFAS")

FPAS_MZ_MAP = {}
def read_database():
    files = [x for x in os.listdir(os.path.dirname(__file__)) if x.endswith('database.txt')]
    
    for file in files:
        print("read database:", file)
        lines = open(os.path.join(os.path.dirname(__file__),file)).read().split('\n')
        for line in lines:
            arr = line.split('\t')
            if len(arr) == 2 and re.match(r'^\d*(\.\d*)?$', arr[1]):
                FPAS_MZ_MAP[round(float(arr[1]),PRECISION)] = arr[0]

def readMongoTask(collection, task_id):
    query = {"_id": ObjectId(task_id)}
    doc = collection.find(query)
    for x in doc:
        return x

def readMongoFromIni():
    global MONGO_HOST,MONGO_PORT,MONGO_DATABASE,MONGO_COLLECTION,COMMOM_FILE_PATH

    import configparser
    config = configparser.ConfigParser()
    config.read(os.path.join(os.path.dirname(__file__), 'config.ini'))
    section = config['DEFAULT']
    MONGO_HOST = section.get('MONGO_HOST')
    MONGO_PORT = section.get('MONGO_PORT')
    MONGO_DATABASE = section.get('MONGO_DATABASE')
    MONGO_COLLECTION = section.get('MONGO_COLLECTION')
    COMMOM_FILE_PATH = section.get('COMMOM_FILE_PATH')

    print(MONGO_HOST,MONGO_PORT,MONGO_DATABASE,MONGO_COLLECTION,COMMOM_FILE_PATH)


def updateStatus(collection, task_id, status, file_key, step, _status):
    query = {"_id": ObjectId(task_id)}

    status[file_key]['step'] = step #{"$numberInt": str(step)}
    status[file_key]['status'] = _status # {"$numberInt": str(_status)}

    collection.update_one(query,  {"$set":{"status":status}})


def readIni(task):
    global TASK_NAME
    global EMAIL

    global KMD_WIDTH
    global MZ_WIDTH
    global PRECISION
    global PRECISION_APPEAR_IN_ALL_SAMPLES
    global INTENSITY_LIST
    global RAW_FILES
    if task is None:
        sys.exit(-1)
    TASK_NAME = task.get('task_name')
    EMAIL = task.get('email')

    KMD_WIDTH = float(task.get('kmd_width'))
    MZ_WIDTH = float(task.get('mz_width'))
    PRECISION = int(task.get('precision'))
    PRECISION_APPEAR_IN_ALL_SAMPLES = int(task.get('precision_appear_in_all_samples'))
    INTENSITY_LIST = [int(x) for x in task.get('intensity_list')]
    RAW_FILES = [x for x in task.get('files') if not (x.endswith('/database.xlsx') or x.endswith('/database.xls'))]


    print("=======CONFIG=======")
    print("TASK_NAME:%s" % TASK_NAME)
    print("EMAIL:%s" % EMAIL)
    print("KMD_WIDTH:%s" % KMD_WIDTH)
    print("MZ_WIDTH:%s" % MZ_WIDTH)
    print("PRECISION:%s" % PRECISION)
    print("PRECISION_APPEAR_IN_ALL_SAMPLES:%s" % PRECISION_APPEAR_IN_ALL_SAMPLES)
    print("INTENSITY_LIST:%s" % str(INTENSITY_LIST))
    print("RAW_FILES:%s" % ",".join(RAW_FILES))
    print("=======CONFIG=======")

def LOG_MSG(*msg):
    tims_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
    print("[%s][INFO] %s" % (tims_str," ".join(msg)))

def LOG_ERROR_MSG(*msg):
    tims_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
    print("[%s][ERROR] %s" % (tims_str," ".join(msg)))
    raise Exception(" ".join(msg))

def create_folder(folder):
    try:
        os.makedirs(folder)
    except FileExistsError:
        pass
def delete_content(folder):
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            LOG_ERROR_MSG('[ERROR]Failed to delete %s. Reason: %s' % (file_path, e))
            

def calculate_kmd(mz):
    km = mz * 50 / 49.99681
    kmd = round(km, 0) - km
    return kmd

def read_raw_data(wb, ws, output_dir):
    # wb = openpyxl.load_workbook(filepath)
    # ws = wb.active
    # if len(wb.sheetnames) != 2 or 'sample' not in wb.sheetnames or 'blank' not in wb.sheetnames:
    #     LOG_ERROR_MSG("ERROR: only 2 sheets allowed. one is named by 'sample', another is 'blank' !")
        
    sample_ws = wb['sample']
    blank_ws = wb['blank']
    def find_in_db(x):
        list_result = []
        for i in [0, -1, +1]:
            tmp_x = round(x + i*1.00794, PRECISION);
            name = FPAS_MZ_MAP.get(tmp_x, None)
            if name is not None:
                flag = i;
                if i == 1:
                    flag = "+1"
                list_result.append(name+ "(%s)" % flag)
        return ",".join(list_result)

    def read_sheet_by_column(ws):
        def is_number(s):
            try:
                float(s)
                return True
            except:
                return False
        if ws.max_column % 2 != 0:
            LOG_ERROR_MSG("ERROR: sheet %s seems like have odd columns, check if have blank cloumn")
            
        c_nums = 0
        c1 = None
        mp_sample = {}
        for cols in ws.iter_cols(min_row = 1, max_col = ws.max_column, max_row = ws.max_row, values_only = True):
            c2 = cols
            # print("before:",len(c2))
            while len(c2) > 0 and c2[-1] is None:
                c2 = c2[:-1]
            # print("after:",len(c2))
            if len(c2) == 0:
                break;
            c_nums = c_nums + 1
            
            if c_nums % 2 == 1:
                c1 = c2
            else:
                if len(c1) != len(c2):
                    LOG_ERROR_MSG("ERROR: data read error")
                name = c1[0].replace(":", "：")
                kmd = [calculate_kmd(x) for x in c1[1:]]
                indb = [find_in_db(x) for x in c1[1:]]
                tmp_result = list(zip(c1[1:],c2[1:], kmd, indb))
                max_mz = -1
                if is_number(c2[0]):
                    max_mz = float(c2[0])
                    if max_mz > 0:
                        tmp_result = [x for x in tmp_result if x[0] <= max_mz]
                mp_sample[name] = tmp_result
        return mp_sample
            
    def sperate_save_sample(mp, location):
        for key in mp:
            __wb = openpyxl.Workbook()
            __ws = __wb.active
            __ws.append(HEAD_TITLE)
            for row in mp[key]:
                __ws.append(row)
            tmp_path = os.path.join(location, key + ".xlsx");
            __wb.save(tmp_path)
            __wb.close()
            LOG_MSG("save " + tmp_path)
            

    blank_mp = read_sheet_by_column(blank_ws)
    sample_mp = read_sheet_by_column(sample_ws)
    wb.close()

    sperate_save_sample(sample_mp, os.path.join(output_dir, DIR_SAMPLE))
    sperate_save_sample(blank_mp, os.path.join(output_dir, DIR_BLANK))
    
    return blank_mp,sample_mp

def save_excel(list_rows, filepath):
    __wb = openpyxl.Workbook()
    __ws = __wb.active
    __ws.append(HEAD_TITLE)
    for row in list_rows:
        __ws.append(row)
    __wb.save(filepath)
    __wb.close()

def filter_with_blank(blank_mp, sample_mp, output_dir):
    mz_filter_set = set()
    for key in blank_mp:
        for row in blank_mp[key]:
            mz_filter_set.add(round(row[0], PRECISION))

    for key in sample_mp:
        tmp_list_row = [x for x in sample_mp[key] if round(x[0], PRECISION) not in mz_filter_set]
        sample_mp[key] = tmp_list_row
        tmp_path = os.path.join(os.path.join(output_dir, DIR_FILTERED_SAMPLE), key + ".xlsx");
        save_excel(tmp_list_row, tmp_path)
        LOG_MSG("save " + tmp_path)

def filter_with_intensity_list(sample_mp, output_dir):
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.chart.legend import LegendEntry
    for key in sample_mp:
        tmp_path = os.path.join(os.path.join(output_dir, DIR_FILTRERD_INTENSITY_LIST), key + ".xlsx");
        __wb = openpyxl.Workbook()
        first_ws = __wb.active
        first_ws.title = "charts"
        for idx,itsy in enumerate(INTENSITY_LIST):
            intensity = int(itsy)
            tmp_list_row = [x for x in sample_mp[key] if round(x[1], PRECISION) >= intensity]
            __ws = __wb.create_sheet(get_proper_name(intensity))
            __ws.append(HEAD_TITLE)
            for row in tmp_list_row:
                __ws.append(row)
            xdata = Reference(__ws, min_col=1, max_col=1, min_row=2, max_row=len(tmp_list_row) + 1)
            ydata = Reference(__ws, min_col=2, max_col=2, min_row=2, max_row=len(tmp_list_row) + 1)
            series1=Series(ydata,xdata,title_from_data=False)

            series1.marker.symbol = "circle"
            series1.graphicalProperties.line.noFill = True
            series1.title = None
            chart = ScatterChart()
            chart.title = "Intensity %s" % get_proper_name(intensity)
            chart.y_axis.title=None
            chart.x_axis.title=None
            chart.append(series1)

            chart.legend = None #hide the legend
            
            pos = chr((idx%3)*9+ord('B')) + str((idx//3)*15 + 2)
            first_ws.add_chart(chart, pos)


        __wb.save(tmp_path)
        __wb.close()
        LOG_MSG("save " + tmp_path)


def init_ouputdir(output_dir):
    print("init folder:",output_dir)
    create_folder(output_dir)
    delete_content(output_dir)
    create_folder(os.path.join(output_dir, DIR_SAMPLE))
    create_folder(os.path.join(output_dir, DIR_BLANK))
    create_folder(os.path.join(output_dir, DIR_FILTERED_SAMPLE))
    
    create_folder(os.path.join(output_dir, DIR_FILTRERD_INTENSITY_LIST))

    create_folder(os.path.join(output_dir, DIR_RESULT))

def find_similar_rows(list_row, index, kmd_low, kmd_high):
    row_set = set()
    list_result_remainder = []
    while index < len(list_row):
        if kmd_low <= list_row[index][2] <= kmd_high:
            # list_result_row.append(list_row[index])
            list_result_remainder.append((list_row[index][0] % 50, index))
            row_set.add(index)
        else:
            break;
        index = index + 1
    list_result_remainder.sort(key=itemgetter(0)) #mz
    return list_result_remainder,row_set

def find_mz_set_of_index(list_tp, index, low, high):
    list_rs = []
    list_debug = []
    while index < len(list_tp):
        if low <= list_tp[index][0] <= high:
            list_rs.append(list_tp[index][1])
            # list_debug.append(list_tp[index])
        else:
            break;
        index = index + 1
    return set(list_rs),list_debug


def make_segment_remainder(list_reminder, remainder_length):
    seg_list = []
    i = 1
    tmp_list = [list_reminder[0],]
    while i < remainder_length:
        c = list_reminder[i][0] - list_reminder[i-1][0]
        if c > MZ_WIDTH:
            if len(tmp_list) >= 3:
                seg_list.append(tmp_list)
            tmp_list = [list_reminder[i],]
        else:
            tmp_list.append(list_reminder[i])
        i = i + 1;
    if len(tmp_list) >= 3:
        seg_list.append(tmp_list)
    return seg_list

def check_result_set_exists(list_set, target_set):
    for s in list_set:
        if target_set <= s:
            return True
    return False
def remove_small_set(list_set, target_set):
    list_rm = [];
    for s in list_set:
        if s < target_set:
            list_rm.append(s)
    for s in list_rm:
        list_set.remove(s)

def get_proper_name(intensity):
    intensity = int(intensity)
    if intensity >= 10000:
        if(intensity%10000) == 0:
            return "%sw" % (intensity//10000)
        else:
            return "%sw" % (intensity/10000)
    else:
        return str(intensity)

def group_kmd_mz(sample_mp, output_dir):
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.chart.legend import LegendEntry
    summary_path = os.path.join(os.path.join(output_dir, DIR_RESULT), "Summary.xlsx");
    summary_wb = openpyxl.Workbook()
    summary_count_ws = summary_wb.active
    summary_count_ws.title = "Group Count"
    summary_count_ws["A1"] = "Intensity"
    for idx,itsy in enumerate(INTENSITY_LIST):
        summary_count_ws["A"+str(idx+2)] = get_proper_name(itsy)

    all_sample_mz_map = {} # intensity: []
    for key_idx,key in enumerate(sample_mp):
        tmp_path = os.path.join(os.path.join(output_dir, DIR_RESULT), key + ".xlsx");
        __wb = openpyxl.Workbook()
        first_ws = __wb.active
        first_ws.title = "charts"
        first_ws.append(["Intensity","Number of group"])
        summary_count_ws.cell(row=1,column=key_idx+2).value = key

        for idx,itsy in enumerate(INTENSITY_LIST):
            single_mz_map = all_sample_mz_map.get(itsy, {})
            all_sample_mz_map[itsy] = single_mz_map
            intensity = int(itsy)
            list_row = [x for x in sample_mp[key] if round(x[1], PRECISION) >= intensity]
            result_list_of_set = find_groups(list_row)
            first_ws["A"+str(idx+2)] = get_proper_name(intensity)
            if len(result_list_of_set) > 0:
                first_ws["B"+str(idx+2)] = len(result_list_of_set)
                summary_count_ws.cell(row=idx+2,column=key_idx+2).value = len(result_list_of_set)

            __ws = __wb.create_sheet(get_proper_name(intensity))
            __ws.append((None,) + HEAD_TITLE)
            if len(result_list_of_set) == 0:
                continue
            max_row_count = 1
            abs_axis = 0.3
            single_mz_set = set()
            for idx2,result_set in enumerate(result_list_of_set):
                __ws.append(("No.%s" % (idx2+1),))
                max_row_count = max_row_count + 1 + len(result_set)
                for row_index in result_set:
                    __ws.append((None,) + list_row[row_index])
                    single_mz_set.add(round(list_row[row_index][0], PRECISION_APPEAR_IN_ALL_SAMPLES))
                    while abs(list_row[row_index][2]) > abs_axis:
                        abs_axis += 0.1
            for mz in single_mz_set:
                if mz in single_mz_map:
                    single_mz_map[mz] = (mz, single_mz_map[mz][1] + 1) + ((single_mz_map[mz][2] + ", " + key),)
                else:
                    single_mz_map[mz] = (mz, 1, key)

            xdata = Reference(__ws, min_col=2, max_col=2, min_row=2, max_row=max_row_count)
            ydata = Reference(__ws, min_col=4, max_col=4, min_row=2, max_row=max_row_count)
            series1=Series(ydata,xdata,title_from_data=False)

            series1.marker.symbol = "circle"
            series1.graphicalProperties.line.noFill = True
            series1.title = None
            chart = ScatterChart()
            chart.title = "%s" % get_proper_name(intensity)
            chart.y_axis.title=None
            chart.x_axis.title=None
            chart.append(series1)

            chart.y_axis.scaling.min = -abs_axis
            chart.y_axis.scaling.max = abs_axis

            chart.legend = None #hide the legend
            
            pos = chr((idx%3)*9+ord('D')) + str((idx//3)*15 + 2)
            first_ws.add_chart(chart, pos)
        __wb.save(tmp_path)
        __wb.close()
        LOG_MSG("save " + tmp_path)


    for idx,itsy in enumerate(INTENSITY_LIST):
        summary_intensity_ws = summary_wb.create_sheet(get_proper_name(itsy))
        single_mz_map = all_sample_mz_map.get(itsy, {})
        single_mz_list = list(single_mz_map.values())
        single_mz_list.sort(key=itemgetter(1, 0),reverse=True) #(mz, count) count ->
        summary_intensity_ws.append(("m/z", "cout", "samples"))
        for data_tuple in single_mz_list:
            summary_intensity_ws.append(data_tuple)

    summary_wb.save(summary_path)
    summary_wb.close()
        




def find_groups(list_row):
    list_row.sort(key=itemgetter(2)) # (mz, intensity, kmd) kmd -> 
    result_list_of_set = []
    last_row_indexset = set()
    for idx,row in enumerate(list_row):
        
        kmd_low, kmd_high = row[2], row[2] + KMD_WIDTH
        list_tuple_remainder,row_set = find_similar_rows(list_row, idx, kmd_low, kmd_high)
        
        if row_set <= last_row_indexset or len(list_tuple_remainder) < 3:
            last_row_indexset = row_set
            continue
        last_row_indexset = row_set

        remainder_length = len(list_tuple_remainder);
        if list_tuple_remainder[-1][0] + MZ_WIDTH > 50:
            flag_index = 0;
            for idx2,row2 in enumerate(list_tuple_remainder):
                if row2[0] <= MZ_WIDTH:
                    flag_index = idx2 + 1
                else:
                    break;
            if flag_index > 0:
                for tp in list_tuple_remainder[0:flag_index]:
                    list_tuple_remainder.append((tp[0]+50, tp[1]))

        segment_remainder_list = make_segment_remainder(list_tuple_remainder,remainder_length)
        for seg in segment_remainder_list:
            if(len(seg) < 3):
                continue
            if(seg[-1][0] - seg[0][0] <= MZ_WIDTH):
                set_tmp = set([x[1] for x in seg])
                # print(set_tmp)
                if check_result_set_exists(result_list_of_set, set_tmp) == False:
                    remove_small_set(result_list_of_set, set_tmp)
                    result_list_of_set.append(set_tmp)

            else:
                for idx_seg,x in enumerate(seg):
                    #print("seg:",seg)
                    mz_low,mz_high = x[0],x[0]+MZ_WIDTH
                    result_set,list_debug = find_mz_set_of_index(seg,idx_seg,mz_low,mz_high)
                    # print("result_set#####:",result_set)
                    if len(result_set) > 3 and check_result_set_exists(result_list_of_set, result_set) == False:
                        remove_small_set(result_list_of_set, result_set)
                        result_list_of_set.append(result_set)

    return result_list_of_set



def createZip(password, zip_name, folder):
    if password is not None and len(password) > 0:
        secret_password = password.encode('utf-8')

        with pyzipper.AESZipFile(zip_name,
                                 'w',
                                 compression=pyzipper.ZIP_LZMA,
                                 encryption=pyzipper.WZ_AES) as zf:
            def getFileList(path):
                files = os.listdir(path)
                for file in files:
                    file_path = os.path.join(path, file)
                    if os.path.isdir(file_path):
                        getFileList(file_path)
                    elif os.path.isfile(file_path):
                        zf.write(file_path)

            zf.setpassword(secret_password)
            getFileList(folder)
    else:
        if(zip_name.endswith('.zip')):
            zip_name = zip_name[:-4]
        shutil.make_archive(zip_name, 'zip', folder) 




if __name__ == '__main__':
    task_id = sys.argv[1];
    sub_path = time.strftime("%Y%m%d_", time.localtime(time.time()))+ task_id;
    workpath = os.path.join(os.path.dirname(__file__), '../data-output/' + sub_path)
    print("workpath:",workpath)
    create_folder(workpath)
    os.chdir(workpath)
    
    readMongoFromIni()
    client = MongoClient(MONGO_HOST, int(MONGO_PORT))
    db = client[MONGO_DATABASE]
    collection = db[MONGO_COLLECTION]
    task = readMongoTask(collection, task_id)

    readIni(task) #read info from mongo
    read_database()

    for x in RAW_FILES:
        file_key = x.replace('.','_')
        # 1. create relevent folder 
        step = 0;
        try:
            if x.endswith('.xls'):
                output_dir = os.path.basename(x)[:-4]
            else:
                output_dir = os.path.basename(x)[:-5]
            # print("output_dir:",output_dir)
            init_ouputdir(output_dir)
            # 2. read data
            print("read file:",x)
            updateStatus(collection, task_id, task['status'], file_key, 1, 0);step = 1
            wb = openpyxl.load_workbook(os.path.join(COMMOM_FILE_PATH,x))
            ws = wb.active
            if len(wb.sheetnames) != 2 or 'sample' not in wb.sheetnames or 'blank' not in wb.sheetnames:
                LOG_ERROR_MSG("ERROR: only 2 sheets allowed. one is named by 'sample', another is 'blank' !")
            updateStatus(collection, task_id, task['status'], file_key, 3, 0);step = 2
            blank_mp,sample_mp = read_raw_data(wb, ws, output_dir)
            updateStatus(collection, task_id, task['status'], file_key, 3, 0);step = 3
            # 3. filter
            filter_with_blank(blank_mp, sample_mp, output_dir)
            updateStatus(collection, task_id, task['status'], file_key, 4, 0);step = 4
            filter_with_intensity_list(sample_mp, output_dir)
            updateStatus(collection, task_id, task['status'], file_key, 5, 0);step = 5
            # 4. find groups 
            group_kmd_mz(sample_mp, output_dir)
            updateStatus(collection, task_id, task['status'], file_key, 6, 0);step = 6
        except:
            updateStatus(collection, task_id, task['status'], file_key, step, 1)
        
    


    

    zip_path = os.path.join(os.path.dirname(__file__), '../data-output/' +sub_path+'_zip')
    create_folder(zip_path)
    zipfile_name = re.subn('[^0-9A-Za-z]','_',task.get('task_name'))[0] + '.zip'
    zipfile_path = os.path.join(zip_path,zipfile_name)

    createZip(task["email"], zipfile_path, '.')

    query = {"_id": ObjectId(task_id)}



    collection.update_one(query,  {"$set":{"result":{"msg":"succ", "file":sub_path+'_zip/'+zipfile_name}}})


    client.close()


    sys.exit()





